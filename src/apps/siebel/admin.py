# coding-style: https://docs.djangoproject.com/en/dev/internals/contributing/writing-code/coding-style/#model-style

from __future__ import unicode_literals

import csv
import ast
from apps.paywall.models import Operation, FinancialTransaction, PaymentProfile, Payment, Subscription
from datetime import date, timedelta, datetime
from django.conf import settings
from django.contrib import admin
from django.http import HttpResponse
from django.urls import reverse
from django.utils import formats, timezone
from django.utils.encoding import smart_str
from django.utils.html import format_html
from django.utils.timezone import get_default_timezone
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from apps.paywall.arc_clients import SalesClient
from apps.paywall.utils_siebel import SiebelConciliationSender
from sentry_sdk import capture_exception
from .models import Logs, Rate, SiebelSubscription, SiebelAction, SiebelLog, LoadTransactionsIdSiebel, \
    PendingSendSiebel, SiebelConfirmationPayment, LogSiebelOvPE, LogSiebelConciliacionPE, LogSiebelOv, LogSiebelClient, \
    LogSiebelConciliacion, LoadProfile, LogSiebelOvPiano, LogUnsubscribePiano, ArcUnsubscribe, LogRenovationPiano, \
    LogSiebelPaymentPiano
from import_export.admin import ExportMixin
from .resources import LogRenovationPianoResource
from rangefilter.filter import DateTimeRangeFilter


def local_format(_date):
    try:
        if not isinstance(_date, datetime):
            return ''

        _date = _date.astimezone(
            get_default_timezone()
        )

        _date = _date.replace(tzinfo=None)

        return _date
    except Exception as e:
        return ''
    except SystemExit:
        return ''


class FirstPaymentFilter(admin.SimpleListFilter):
    title = 'FirstPayment'
    parameter_name = 'first_payment'

    def lookups(self, request, model_admin):
        return (
            ('1', 'Primer Pago'),
            ('2', 'Recurrencias'),
        )

    def queryset(self, request, queryset):
        if self.value() == '1':
            return queryset.filter(num_liquidacion='VENTA')
        elif self.value() == '2':
            return queryset.exclude(num_liquidacion='VENTA')
        else:
            return queryset


class OperationEmptyFilter(admin.SimpleListFilter):
    title = 'OperationEmpty'
    parameter_name = 'pperation_empty'

    def lookups(self, request, model_admin):
        return (
            ('1', 'Sin Operacion'),
        )

    def queryset(self, request, queryset):
        if self.value() == '1':
            return queryset.filter(operation__isnull=True)
        else:
            return queryset


class CIPFilter(admin.SimpleListFilter):
    title = 'cipfilter'
    parameter_name = 'cip_filter'

    def lookups(self, request, model_admin):
        return (
            ('1', 'Sin CIP'),
            ('2', 'Con CIP'),
        )

    def queryset(self, request, queryset):
        if self.value() == '1':
            return queryset.filter(cip__isnull=True)
        elif self.value() == '2':
            return queryset.filter(cip__isnull=False)
        else:
            return queryset


@admin.register(SiebelConfirmationPayment)
class SiebelConfirmationPaymentAdmin(admin.ModelAdmin):
    list_display = ['get_plan', 'get_data', 'get_extra']
    search_fields = ('cod_delivery', 'num_liquidacion',)
    readonly_fields = (
        'operation',
    )
    list_filter = (OperationEmptyFilter, CIPFilter, FirstPaymentFilter,)

    def get_plan(self, obj):
        if obj.operation:
            return format_html(
                '<strong>{name_plan} </strong></br>'
                '<i class="fas fa-key"></i> ID {subscription_id}</br>'
                '<i class="fas fa-newspaper"></i> {brand}</br>'
                '<b>Estado:</b> {estado}</br>',
                name_plan=obj.operation.plan.plan_name if obj.operation else '',
                brand=obj.operation.payment.partner.partner_name if obj.operation else '',
                subscription_id=obj.operation.payment.subscription.arc_id if obj.operation else '',
                estado=obj.operation.payment.subscription.get_state_display() if obj.operation else ''
            )
        elif obj.cip:
            return format_html(
                '<strong>{name_plan} </strong></br>'
                '<i class="fas fa-key"></i> ID {subscription_id}</br>'
                '<i class="fas fa-newspaper"></i> {brand}</br>'
                '<b>Estado:</b> {estado}</br>',
                name_plan=obj.cip.plan.plan_name,
                brand=obj.cip.subscription.partner.partner_name,
                subscription_id=obj.cip.subscription.arc_id,
                estado=obj.cip.subscription.get_state_display()
            )
        else:
            return ''

    def get_data(self, obj):
        tz = timezone.get_current_timezone()
        if obj.fecha_de_emision:
            tz_fecha_de_emision = obj.fecha_de_emision.astimezone(tz)
            fecha_emision = formats.date_format(tz_fecha_de_emision, settings.DATETIME_FORMAT)
        else:
            fecha_emision = None

        return format_html(
            '<b>ENTECODE:</b> {code_ente}</br>'
            '<b>Delivery:</b> {cod_delivery} </br>'
            '<b>Nro renovacion:</b> {nro_renovacion} </br>'
            '<b>Fecha de Emision:</b> {fecha_de_emision} </br>'
            '<b>Fecha de Creación:</b> {creacion}</br>'
            '<b>Fecha de Actualización:</b> {update_}</br>'
            '<b>operation: </b> {operation}</br>'
            '<b>cip: </b> {cip}',
            code_ente=obj.code_ente if obj.code_ente else '',
            cod_delivery=obj.cod_delivery if obj.cod_delivery else '',
            nro_renovacion=obj.nro_renovacion if obj.nro_renovacion else '',
            fecha_de_emision=fecha_emision,
            creacion=formats.date_format(obj.created.astimezone(tz), settings.DATETIME_FORMAT),
            update_=formats.date_format(obj.last_updated.astimezone(tz), settings.DATETIME_FORMAT),
            operation=obj.operation,
            cip=obj.cip
        )

    def get_extra(self, obj):
        return format_html(
            '<b>Cod interno Comprobante:</b> {cod_interno_comprobante} </br>'
            '<b>Folio SUNAT:</b> {folio_sunat} </br>'
            '<b>Monto:</b> {monto} </br>'
            '<b>Nro Liquidación:</b> {num_liquidacion}',
            cod_interno_comprobante=obj.cod_interno_comprobante if obj.cod_interno_comprobante else '',
            folio_sunat=obj.folio_sunat if obj.folio_sunat else '',
            monto=obj.monto if obj.monto else '',
            num_liquidacion=obj.num_liquidacion if obj.num_liquidacion else ''

        )


@admin.register(Logs)
class LogsAdmin(admin.ModelAdmin):
    list_display = ['id', 'delivery', 'log_type', 'created', 'state']


@admin.register(Rate)
class RateAdmin(admin.ModelAdmin):
    search_fields = ('siebel_code_promo',)
    list_display = ['plan', 'rate_name', 'siebel_code_promo', 'rate_neto', 'rate_igv', 'rate_total', 'created',
                    'siebel_id', 'siebel_code',
                    'state', 'type']
    list_filter = ('plan__partner__partner_name', 'plan',)


@admin.register(SiebelSubscription)
class SiebelSubscriptionAdmin(admin.ModelAdmin):
    list_display = ['subscription', 'siebel_delivery', 'created', 'date_low', 'last_updated']
    search_fields = ('siebel_delivery', 'subscription__arc_id',)

    def date_low(self, obj):
        tz = timezone.get_current_timezone()
        if obj.subscription.date_anulled:
            tz_fecha_de_baja = obj.subscription.date_anulled.astimezone(tz)
            return formats.date_format(tz_fecha_de_baja, settings.DATETIME_FORMAT)
        return ''


@admin.register(SiebelAction)
class SiebelActionAdmin(admin.ModelAdmin):
    pass


@admin.register(SiebelLog)
class SiebelLogAdmin(admin.ModelAdmin):
    search_fields = ('url', 'request_text',)
    list_display = ['url']


def export_csv(modeladmin, request, queryset):
    import csv
    fecha_report = date.today().strftime("%d-%m-%Y")
    from django.utils.encoding import smart_str
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=report' + fecha_report + '.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8'))  # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str("transaction_id"),
        smart_str("cliente"),
        smart_str(u"ov"),
        smart_str(u"conciliacion"),
        smart_str(u"log response"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.transaction_id),
            smart_str(obj.cliente),
            smart_str(obj.ov),
            smart_str(obj.conciliation),
            smart_str(obj.log_response),
        ])
    return response


export_csv.short_description = u"Export CSV"


@admin.register(PendingSendSiebel)
class PendingSendSiebelAdmin(admin.ModelAdmin):
    list_display = ('transaction_id',)
    actions = [export_csv]


def get_transaction_id(obj):
    try:
        # transaction_id = obj.payment.payment_financial_transaction.transaction_id
        payu_ope = FinancialTransaction.objects.get(
            order_number=obj.payment.arc_order
        )
        transaction_id = payu_ope.transaction_id
    except Exception:
        transaction_id = ''
    return transaction_id


def get_fecha_pago(obj):
    try:
        fecha_pago = local_format(obj.payment.date_payment)
    except Exception as e:
        fecha_pago = ''
    return fecha_pago


def get_first_payment(obj):
    """
    first_payment = 'No es el primer pago'
    operation_obj = Operation.objects.filter(
        payment__subscription__arc_id=obj.payment.subscription.arc_id,
        ope_amount__gte=5
    ).order_by('payment__transaction_date').first()
    if operation_obj.payment.payu_transaction and obj.payment.payu_transaction:
        if operation_obj.payment.payu_transaction == obj.payment.payu_transaction:
            first_payment = 'Primer pago'
    return first_payment
    """
    return ''


def get_arc_id(obj):
    try:
        arc_id = str(obj.payment.subscription.arc_id)
    except Exception as e:
        arc_id = ''
    return arc_id


def get_last_send_client(obj):
    try:
        log_cl = LogSiebelClient.objects.filter(payment_profile=obj).last()
        return local_format(log_cl.created)
    except Exception as e:
        return ''


def get_last_send_ov(obj):
    try:
        log_ov = LogSiebelOv.objects.filter(operation=obj).last()
        return local_format(log_ov.created)
    except Exception as e:
        return ''


def get_payment_confirmation(obj):
    if SiebelConfirmationPayment.objects.filter(operation=obj).exists():
        return 'Recepcionado'
    else:
        return 'No recepcionado'


def date_get_payment_confirmation(obj):
    confirmation = SiebelConfirmationPayment.objects.order_by('created').filter(operation=obj).last()
    if confirmation:
        return local_format(confirmation.created)
    else:
        return ''


def get_siebel_entecode(obj):
    try:
        siebel_entecode = obj.payment_profile.siebel_entecode
        siebel_direction = obj.payment_profile.siebel_entedireccion
        if siebel_entecode and siebel_direction:
            return siebel_entecode
        else:
            return ''
    except Exception:
        return ''


def get_doc_num(obj):
    try:
        return obj.payment_profile.prof_doc_num
    except Exception:
        return ''


def get_type_doc(obj):
    try:
        return obj.payment_profile.prof_doc_type
    except Exception:
        return ''


def get_cellphone(obj):
    if obj.payment_profile:
        return '{phone}'.format(phone=obj.payment_profile.prof_phone)
    else:
        return ''


def get_full_name(obj):
    if obj.payment_profile:
        full_name = '{name} {last_name} {last_name_mother}'.format(
            name=obj.payment_profile.prof_name,
            last_name=obj.payment_profile.prof_lastname,
            last_name_mother=obj.payment_profile.prof_lastname_mother)
        return full_name
    else:
        return ''


def last_payment_exist(obj):
    operations = Operation.objects.filter(
        payment__subscription__arc_id=obj.payment.subscription.arc_id,
        payment__date_payment__lte=obj.payment.date_payment
    ).order_by('payment__date_payment')
    for ope in operations:
        if ope.payment.payu_transaction == obj.payment.payu_transaction:
            if last_payment.ope_amount > 3:
                return True
            else:
                return False
        last_payment = ope
    return False


def get_resumen(siebel_entecode, delivery, obj, first_send_payment):
    if not siebel_entecode:
        return 'no se creo entecode'
    elif not delivery:
        return 'no se creo delivery'
    else:
        if obj.conciliation_cod_response == '1':
            return 'Enviado'
    return ''


def get_delivery(obj):
    try:
        delivery = obj.payment.subscription.delivery
    except Exception as e:
        delivery = ''
    return delivery


def get_date_delivery(obj):
    try:
        obj_log = LogSiebelOv.objects.filter(
            operation__payment__subscription=obj.payment.subscription
        ).order_by('created').last()
    except:
        obj_log = ''

    if obj_log:
        return local_format(obj_log.created)
    else:
        return ''


def get_payment_profile(obj):
    try:
        if not obj.payment_profile.siebel_entedireccion:
            ente_envio = obj.payment_profile.siebel_request
            ente_respuesta = obj.payment_profile.siebel_response
        else:
            ente_envio = ''
            ente_respuesta = ''
    except Exception:
        ente_envio = ''
        ente_respuesta = ''

    return ente_envio, ente_respuesta


def get_envio_respuesta_ov(obj, delivery):
    try:
        obj_log = LogSiebelOv.objects.filter(
            operation__payment__subscription=obj.payment.subscription
        ).order_by('created').last()
    except:
        obj_log = ''

    if obj_log:
        ov_pedido = obj_log.log_request
        ov_respuesta = obj_log.log_response
    else:
        ov_pedido = ''
        ov_respuesta = ''
    return ov_pedido, ov_respuesta


def get_monto(obj):
    try:
        monto = obj.payment.pa_amount
    except Exception as e:
        monto = ''
    return monto


def get_estado_pago(obj):
    try:
        if obj.conciliation_cod_response == '1':
            estado_pago = 'Enviado'
        else:
            estado_pago = 'No Enviado'
    except Exception as e:
        estado_pago = 'No Enviado'
    return estado_pago


def first_send_payment_state(obj):
    r_response = obj.recurrencia_response
    if r_response:
        response_dict = ast.literal_eval(r_response)
        response_service = response_dict.get('response', {}).get('respuesta', None)
        if int(response_service) == 1:
            return 'Enviado'
        else:
            return 'Error'
    else:
        return 'No enviado'


def date_first_send_payment(obj):
    r_request = obj.recurrencia_request
    if r_request:
        last_object = LogSiebelConciliacion.objects.filter(operation=obj).exclude(log_recurrence_request__isnull=True)\
            .exclude(log_recurrence_request__exact='').order_by('created').last()
        if last_object:
            return local_format(last_object.created)
    return ''


def export_csv_order_to_transactions(modeladmin, request, queryset):
    fecha_report = date.today().strftime("%d-%m-%Y")
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=report_order_transactionid' + fecha_report + '.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8'))
    writer.writerow([
        smart_str(u"OrderId"),
        smart_str(u"TransactionId"),
    ])
    for obj in queryset:
        transactions = obj.transaction_id
        list_orders = transactions.splitlines()

        financial_transactions = FinancialTransaction.objects.filter(
            order_id__in=list_orders
        )
        for transaction in financial_transactions:
            row = [
                transaction.order_id or '',
                transaction.transaction_id or ''
            ]
            writer.writerow(row)
    return response


def send_faltantes(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=log_envio.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8'))

    for obj in queryset:
        transactions = obj.transaction_id
        list_transactions = transactions.splitlines()

        for id_transaction in list_transactions:
            log_envio = ''
            operation_list = Operation.objects.filter(
                ope_amount__gte=5,
                payment__pa_origin='RECURRENCE',
                payment_profile__siebel_entecode__isnull=False,
                payment_profile__siebel_entedireccion__isnull=False,
                payment__subscription__delivery__isnull=False,
                payment__payu_transaction=id_transaction
            )

            operation_list.exclude(recurrencia_response_state=True)
            operation_list = operation_list.order_by('payment__date_payment')

            for operation in operation_list:
                if not FinancialTransaction.objects.filter(order_number=operation.payment.arc_order,
                                                           transaction_type='Refund').exists():
                    siebel_client = SiebelConciliationSender(operation, False)
                    try:
                        log_envio = 'Iniciando envio: {operation_id}'.format(operation_id=operation.id)
                        siebel_client.send_payment_faltantes()
                    except Exception:
                        capture_exception()

                    print('Termino la ejecucion del comando')

            writer.writerow([
                id_transaction,
                log_envio,
                operation_list
            ])
    return response


def export_csv_transactions_to_orders(modeladmin, request, queryset):
    list_t = []
    fecha_report = date.today().strftime("%d-%m-%Y")
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=report_order_transactionid' + fecha_report + '.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8'))
    writer.writerow([
        smart_str(u"OrderId"),
        smart_str(u"TransactionId"),
    ])
    for obj in queryset:
        transactions = obj.transaction_id
        list_transactions = transactions.splitlines()

        financial_transactions = FinancialTransaction.objects.filter(
            transaction_id__in=list_transactions
        )
        for transaction in financial_transactions:
            row = [
                transaction.order_id or '',
                transaction.transaction_id or ''
            ]
            writer.writerow(row)
    return response


def export_csv_anticipos_by_order_id(modeladmin, request, queryset):
    list_t = []
    fecha_report = date.today().strftime("%d-%m-%Y")
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=report_anticipos_' + fecha_report + '.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8'))
    writer.writerow([
        smart_str(u"Delivery"),
        smart_str(u"fecha de baja"),
        smart_str(u"Estado"),
        smart_str(u"TransactionId"),
    ])
    for obj in queryset:
        transactions = obj.transaction_id
        list_transactions = transactions.splitlines()

        financial_transactions = FinancialTransaction.objects.filter(
            order_id__in=list_transactions
        )
        for transaction in financial_transactions:
            list_t.append(transaction.transaction_id)
        users = Operation.objects.filter(
            payment__payu_transaction__in=list_t
        )

        for obj in users:
            try:
                fecha_baja = local_format(obj.payment.subscription.date_anulled)
            except Exception as e:
                fecha_baja = ''

            row = [
                obj.payment.subscription.delivery or '',
                fecha_baja,
                obj.payment.subscription.get_state_display() or '',
                obj.payment.payu_transaction
            ]
            writer.writerow(row)
    return response


def export_csv_anticipos_by_transaction_id(modeladmin, request, queryset):
    fecha_report = date.today().strftime("%d-%m-%Y")
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=report_anticipos_' + fecha_report + '.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8'))
    writer.writerow([
        smart_str(u"Delivery"),
        smart_str(u"fecha de baja"),
        smart_str(u"Estado"),
        smart_str(u"TransactionId"),
    ])
    for obj in queryset:
        transactions = obj.transaction_id
        list_transactions = transactions.splitlines()

        users = Operation.objects.filter(
            payment__payu_transaction__in=list_transactions
        )

        for obj in users:
            try:
                fecha_baja = local_format(obj.payment.subscription.date_anulled)
            except Exception as e:
                fecha_baja = ''

            row = [
                obj.payment.subscription.delivery or '',
                fecha_baja,
                obj.payment.subscription.get_state_display() or '',
                obj.payment.payu_transaction
            ]
            writer.writerow(row)
    return response


def export_csv_anticipos_by_delivery(modeladmin, request, queryset):
    fecha_report = date.today().strftime("%d-%m-%Y")
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=report_anticipos_' + fecha_report + '.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8'))
    writer.writerow([
        smart_str(u"Delivery"),
        smart_str(u"fecha de baja"),
        smart_str(u"Estado"),
        smart_str(u"TransactionIds"),
    ])
    for obj in queryset:
        item_delivery = obj.transaction_id
        list_delivery = item_delivery.splitlines()

        users = Subscription.objects.filter(
            delivery__in=list_delivery
        )

        for obj in users:
            list_payments = ''
            payments = Payment.objects.filter(
                subscription=obj
            )
            for payment in payments:
                if list_payments:
                    list_payments = list_payments + '*' + str(payment.payu_transaction)
                else:
                    list_payments = str(payment.payu_transaction)

            try:
                fecha_baja = local_format(obj.date_anulled)
            except Exception as e:
                fecha_baja = ''

            row = [
                obj.delivery or '',
                fecha_baja,
                obj.get_state_display() or '',
                list_payments
            ]
            writer.writerow(row)
    return response


def export_csv_transactions_detail(modeladmin, request, queryset):
    fecha_report = date.today().strftime("%d-%m-%Y")
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=report_operation_' + fecha_report + '.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8'))
    writer.writerow([
        smart_str(u"Resumen"),
        smart_str(u"Subscription Id"),
        smart_str(u"Trama Envio ENTECODE"),
        smart_str(u"Trama Respuesta ENTECODE"),
        smart_str(u"Delivery"),
        smart_str(u"Trama Envio Ov"),
        smart_str(u"Trama Respuesta Ov"),
        smart_str(u"num_liquida_id - TransactionId PAYU - sent"),
        smart_str(u"Transaction Id"),
        smart_str(u"Monto"),
        smart_str(u"monto_enviado"),
        smart_str(u"Fecha de Pago"),
        smart_str(u"Plan"),
        smart_str(u"Pago enviado a Siebel"),
        smart_str(u"Nombre del promocion siebel"),
        smart_str(u"tipo - (web o recurrence)"),
        smart_str(u"primer pago(diferente de cero)"),
        smart_str(u"ultimo envio a siebel"),
        smart_str(u"confirmacion de pago"),
    ])
    for obj in queryset:
        transactions = obj.transaction_id
        list_transactions = transactions.splitlines()
        users = Operation.objects.filter(
            payment__payu_transaction__in=list_transactions
        )

        for obj in users:
            first_payment = get_first_payment(obj)

            if first_payment != 'Primer pago':
                operations = Operation.objects.filter(
                    payment__subscription__arc_id=obj.payment.subscription.arc_id,
                    ope_amount__gte=5
                ).order_by('created')
                for op in operations:
                    delivery = get_delivery(op)
                    ov_pedido, ov_respuesta = get_envio_respuesta_ov(op, delivery)
                    ente_envio, ente_respuesta = get_payment_profile(op)
                    siebel_entecode = get_siebel_entecode(op)
                    resumen = get_resumen(siebel_entecode, delivery, op, None)
                    transaction_id = get_transaction_id(op)
                    fecha_pago = get_fecha_pago(op)

                    row = [
                        resumen,
                        get_arc_id(op),
                        ente_envio,
                        ente_respuesta,
                        delivery,
                        ov_pedido,
                        ov_respuesta,
                        '',
                        transaction_id,
                        get_monto(op),
                        '',
                        fecha_pago,
                        '',
                        '',
                        '',
                        '',
                        get_first_payment(op),
                        get_last_send_ov(op),
                        get_payment_confirmation(op)
                    ]
                    writer.writerow(row)
            arc_id = get_arc_id(obj)
            delivery = get_delivery(obj)

            try:
                if obj.conciliation_siebel_request:
                    start = '<tem:num_liquida_id>'
                    end = '</tem:num_liquida_id>'
                    csr = obj.conciliation_siebel_request
                    num_liquida_id_sent = csr[csr.find(start) + len(start):csr.find(end)]
                else:
                    num_liquida_id_sent = ''
            except Exception as e:
                num_liquida_id_sent = ''

            monto = get_monto(obj)

            try:
                if obj.conciliation_siebel_request:
                    start = '<tem:monto_cobrado>'
                    end = '</tem:monto_cobrado>'
                    csr = obj.conciliation_siebel_request
                    monto_enviado = csr[csr.find(start) + len(start):csr.find(end)]
                else:
                    monto_enviado = ''
            except Exception as e:
                monto_enviado = ''

            fecha_pago = get_fecha_pago(obj)

            try:
                plan_name = obj.payment.subscription.plan.plan_name
            except Exception as e:
                plan_name = ''

            try:
                if obj.conciliation_cod_response == '1':
                    estado_pago = 'Enviado'
                else:
                    estado_pago = 'No Enviado'
            except Exception as e:
                estado_pago = 'No Enviado'

            try:
                if obj.siebel_request:
                    start = '<eco:ProdPromName>'
                    end = '</eco:ProdPromName>'
                    s = obj.siebel_request
                    prod_prom_name = s[s.find(start) + len(start):s.find(end)]
                else:
                    prod_prom_name = ''
            except Exception as e:
                prod_prom_name = ''

            try:
                if obj.payment.pa_origin == 'WEB':
                    recurrence_value = 'Primera Venta'
                else:
                    recurrence_value = 'Recurrencia'
            except Exception as e:
                recurrence_value = ''

            ov_pedido, ov_respuesta = get_envio_respuesta_ov(obj, delivery)
            ente_envio, ente_respuesta = get_payment_profile(obj)
            siebel_entecode = get_siebel_entecode(obj)
            resumen = get_resumen(siebel_entecode, delivery, obj, None)
            transaction_id = get_transaction_id(obj)

            row = [
                resumen,
                arc_id,
                ente_envio,
                ente_respuesta,
                delivery,
                ov_pedido,
                ov_respuesta,
                num_liquida_id_sent,
                transaction_id,
                monto,
                monto_enviado,
                fecha_pago,
                plan_name,
                estado_pago,
                prod_prom_name,
                recurrence_value,
                first_payment,
                get_last_send_ov(obj),
                get_payment_confirmation(obj)
            ]
            writer.writerow(row)
    return response

export_csv_transactions_detail.short_description = u"Export transacciones CSV - agrupado"


def export_transactions_to_xlsx(modeladmin, request, queryset):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    for obj in queryset:
        transactions = obj.transaction_id
        list_transactions = transactions.splitlines()
        users_queryset = Operation.objects.filter(
            payment__payu_transaction__in=list_transactions
        )

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-transactions.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Transactions'

        # Define the titles for columns
        columns = [
            'Resumen',
            'Subscription Id',
            'Nro de documento',
            'tipo de documento',
            'Trama Envio ENTECODE',
            'Trama Respuesta ENTECODE',
            'Delivery',
            'Ultima fecha de creacion del delivery',
            'Trama Envio Ov',
            'Trama Respuesta Ov',
            'Trama Envio 1er envioPago',
            'Trama Respuesta 1er envioPago',
            'num_liquida_id - TransactionId PAYU - sent',
            'Transaction Id',
            'Monto',
            'monto_enviado',
            'Fecha de Pago',
            'Plan',
            'Pago enviado a Siebel',
            'Nombre del promocion siebel',
            'tipo - (web o recurrence)',
            'primer pago(diferente de cero)',
            'ultimo envio a siebel(Cliente)',
            'ultimo envio a siebel(OV)',
            '1er envio de pago(De Paywall a Siebel - recepción)',
            'Fecha de 1er envio de pago(De Paywall a Siebel - recepción)',
            'confirmacion de pago(De Siebel a Paywall - recepción)',
            'Fecha de confirmacion de pago (De Siebel a Paywall - recepción)',
            'Cargado'
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all movies
        for obj in users_queryset:
            row_num += 1

            # Define the data for each cell in the row
            first_payment = get_first_payment(obj)

            if first_payment != 'Primer pago':
                operations = Operation.objects.filter(
                    payment__subscription__arc_id=obj.payment.subscription.arc_id,
                    ope_amount__gte=5
                ).order_by('created')
                for op in operations:
                    delivery = get_delivery(op)
                    ov_pedido, ov_respuesta = get_envio_respuesta_ov(op, delivery)
                    ente_envio, ente_respuesta = get_payment_profile(op)
                    siebel_entecode = get_siebel_entecode(op)
                    first_send_payment = first_send_payment_state(op)
                    resumen = get_resumen(siebel_entecode, delivery, op, first_send_payment)
                    transaction_id = get_transaction_id(op)
                    fecha_pago = get_fecha_pago(op)

                    row = [
                        resumen,
                        get_arc_id(op),
                        '',
                        '',
                        ente_envio,
                        ente_respuesta,
                        delivery,
                        '',
                        ov_pedido,
                        ov_respuesta,
                        op.recurrencia_request,
                        op.recurrencia_response,
                        '',
                        transaction_id,
                        get_monto(op),
                        '',
                        fecha_pago,
                        '',
                        get_estado_pago(op),
                        '',
                        '',
                        get_first_payment(op),
                        get_last_send_client(op.payment_profile),
                        get_last_send_ov(op),
                        first_send_payment,
                        date_first_send_payment(op),
                        get_payment_confirmation(op),
                        date_get_payment_confirmation(op),
                        ''
                    ]
                    # Assign the data for each cell of the row
                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.fill = PatternFill("solid", fgColor="0000FFFF")
                    row_num += 1
            arc_id = get_arc_id(obj)
            delivery = get_delivery(obj)

            try:
                if obj.conciliation_siebel_request:
                    start = '<tem:num_liquida_id>'
                    end = '</tem:num_liquida_id>'
                    csr = obj.conciliation_siebel_request
                    num_liquida_id_sent = csr[csr.find(start) + len(start):csr.find(end)]
                else:
                    num_liquida_id_sent = ''
            except Exception as e:
                num_liquida_id_sent = ''

            monto = get_monto(obj)

            try:
                if obj.conciliation_siebel_request:
                    start = '<tem:monto_cobrado>'
                    end = '</tem:monto_cobrado>'
                    csr = obj.conciliation_siebel_request
                    monto_enviado = csr[csr.find(start) + len(start):csr.find(end)]
                else:
                    monto_enviado = ''
            except Exception as e:
                monto_enviado = ''

            fecha_pago = get_fecha_pago(obj)

            try:
                plan_name = obj.payment.subscription.plan.plan_name
            except Exception as e:
                plan_name = ''

            try:
                if obj.siebel_request:
                    start = '<eco:ProdPromName>'
                    end = '</eco:ProdPromName>'
                    s = obj.siebel_request
                    prod_prom_name = s[s.find(start) + len(start):s.find(end)]
                else:
                    prod_prom_name = ''
            except Exception as e:
                prod_prom_name = ''

            try:
                if obj.payment.pa_origin == 'WEB':
                    recurrence_value = 'Primera Venta'
                else:
                    recurrence_value = 'Recurrencia'
            except Exception as e:
                recurrence_value = ''

            ov_pedido, ov_respuesta = get_envio_respuesta_ov(obj, delivery)
            ente_envio, ente_respuesta = get_payment_profile(obj)
            siebel_entecode = get_siebel_entecode(obj)
            first_send_payment = first_send_payment_state(obj)
            resumen = get_resumen(siebel_entecode, delivery, obj, first_send_payment)
            transaction_id = get_transaction_id(obj)

            row = [
                resumen,
                arc_id,
                get_doc_num(obj),
                get_type_doc(obj),
                ente_envio,
                ente_respuesta,
                delivery,
                get_date_delivery(obj),
                ov_pedido,
                ov_respuesta,
                obj.recurrencia_request,
                obj.recurrencia_response,
                num_liquida_id_sent,
                transaction_id,
                monto,
                monto_enviado,
                fecha_pago,
                plan_name,
                get_estado_pago(obj),
                prod_prom_name,
                recurrence_value,
                first_payment,
                get_last_send_client(obj.payment_profile),
                get_last_send_ov(obj),
                first_send_payment,
                date_first_send_payment(obj),
                get_payment_confirmation(obj),
                date_get_payment_confirmation(obj),
                'Si'
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response


def export_xls_transactions(modeladmin, request, queryset):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    for obj in queryset:
        transactions = obj.transaction_id
        list_transactions = transactions.splitlines()
        users_queryset = Operation.objects.filter(
            payment__payu_transaction__in=list_transactions
        )

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-transactions-sin_agrupar.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Transactions'

        # Define the titles for columns
        columns = [
            'Resumen',
            'Subscription Id',
            'Nro de documento',
            'Tipo de documento',
            'Nombre',
            'Telefono',
            'Trama Envio ENTECODE',
            'Trama Respuesta ENTECODE',
            'Delivery',
            'Trama Envio Ov',
            'Trama Respuesta Ov',
            'Trama Envio 1er envioPago',
            'Trama Respuesta 1er envioPago',
            'num_liquida_id - TransactionId PAYU - sent',
            'Transaction Id',
            'Monto',
            'monto_enviado',
            'Fecha de Pago',
            'Plan',
            'Pago enviado a Siebel',
            'Nombre del promocion siebel',
            'tipo - (web o recurrence)',
            'primer pago(diferente de cero)',
            'ultimo envio a siebel(Cliente)',
            'ultimo envio a siebel(OV)',
            '1er envio de pago(De Paywall a Siebel - recepción)',
            'Fecha de 1er envio de pago(De Paywall a Siebel - recepción)',
            'confirmacion de pago(De Siebel a Paywall - recepción)',
            'Fecha de confirmacion de pago (De Siebel a Paywall - recepción)',
            'Cargado'
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all movies
        for obj in users_queryset:
            row_num += 1
            # Define the data for each cell in the row
            first_payment = get_first_payment(obj)
            arc_id = get_arc_id(obj)
            delivery = get_delivery(obj)

            try:
                if obj.conciliation_siebel_request:
                    start = '<tem:num_liquida_id>'
                    end = '</tem:num_liquida_id>'
                    csr = obj.conciliation_siebel_request
                    num_liquida_id_sent = csr[csr.find(start) + len(start):csr.find(end)]
                else:
                    num_liquida_id_sent = ''
            except Exception as e:
                num_liquida_id_sent = ''

            monto = get_monto(obj)

            try:
                if obj.conciliation_siebel_request:
                    start = '<tem:monto_cobrado>'
                    end = '</tem:monto_cobrado>'
                    csr = obj.conciliation_siebel_request
                    monto_enviado = csr[csr.find(start) + len(start):csr.find(end)]
                else:
                    monto_enviado = ''
            except Exception as e:
                monto_enviado = ''

            fecha_pago = get_fecha_pago(obj)

            try:
                plan_name = obj.payment.subscription.plan.plan_name
            except Exception as e:
                plan_name = ''

            try:
                if obj.siebel_request:
                    start = '<eco:ProdPromName>'
                    end = '</eco:ProdPromName>'
                    s = obj.siebel_request
                    prod_prom_name = s[s.find(start) + len(start):s.find(end)]
                else:
                    prod_prom_name = ''
            except Exception as e:
                prod_prom_name = ''

            try:
                if obj.payment.pa_origin == 'WEB':
                    recurrence_value = 'Primera Venta'
                else:
                    recurrence_value = 'Recurrencia'
            except Exception as e:
                recurrence_value = ''

            ov_pedido, ov_respuesta = get_envio_respuesta_ov(obj, delivery)
            ente_envio, ente_respuesta = get_payment_profile(obj)
            siebel_entecode = get_siebel_entecode(obj)
            first_send_payment = first_send_payment_state(obj)
            resumen = get_resumen(siebel_entecode, delivery, obj, first_send_payment)
            transaction_id = get_transaction_id(obj)
            document_number = get_doc_num(obj)
            type_document = get_type_doc(obj)
            row = [
                resumen,
                arc_id,
                document_number,
                type_document,
                get_full_name(obj),
                get_cellphone(obj),
                ente_envio,
                ente_respuesta,
                delivery,
                ov_pedido,
                ov_respuesta,
                obj.recurrencia_request,
                obj.recurrencia_response,
                num_liquida_id_sent,
                transaction_id,
                monto,
                monto_enviado,
                fecha_pago,
                plan_name,
                get_estado_pago(obj),
                prod_prom_name,
                recurrence_value,
                first_payment,
                get_last_send_client(obj.payment_profile),
                get_last_send_ov(obj),
                first_send_payment,
                date_first_send_payment(obj),
                get_payment_confirmation(obj),
                date_get_payment_confirmation(obj),
                'Si'
            ]
            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response

export_xls_transactions.short_description = u"Export transacciones xls - sin agrupar"


@admin.register(LoadTransactionsIdSiebel)
class LoadTransactionsIdSiebelAdmin(admin.ModelAdmin):
    list_display = ('transaction_id', 'tipo',)
    actions = [
        export_xls_transactions,
        export_csv_transactions_detail,
        export_transactions_to_xlsx,
        export_csv_anticipos_by_order_id,
        export_csv_anticipos_by_transaction_id,
        export_csv_anticipos_by_delivery,
        export_csv_order_to_transactions,
        export_csv_transactions_to_orders,
        send_faltantes
    ]


@admin.register(LogSiebelOvPE)
class LogSiebelOvPEAdmin(admin.ModelAdmin):
    list_display = ('created', 'get_cip',)
    search_fields = ('log_request', 'log_response', 'cip__cip',)

    def get_cip(self, obj):
        if obj.cip:
            return obj.cip.cip


@admin.register(LogSiebelConciliacionPE)
class LogSiebelConciliacionPEAdmin(admin.ModelAdmin):
    list_display = ('created', 'get_cip',)
    search_fields = ('log_request', 'cip__cip',)

    def get_cip(self, obj):
        if obj.cip:
            return obj.cip.cip


def update_profile(modeladmin, request, queryset):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    lista = []
    for obj in queryset:
        lista.append({
            "id_profile": obj.id_profile,
            "arc_id": obj.arc_id
        })

    for lis in lista:
        profile = PaymentProfile.objects.get(id=lis.get('id_profile'))
        try:
            subscription = Subscription.objects.get(
                arc_id=lis.get('arc_id')
            )
            if subscription:
                subscription.payment_profile = profile
                subscription.save()
        except Exception as e:
            print(e)
            pass

        try:
            payments = Payment.objects.filter(
                subscription__arc_id=lis.get('arc_id')
            )
            for payment in payments:
                payment.payment_profile = profile
                payment.save()
        except Exception as e:
            print(e)
            pass

        try:
            operations = Operation.objects.filter(
                payment__subscription__arc_id=lis.get('arc_id')
            )
            for operation in operations:
                operation.payment_profile = profile
                operation.save()
        except Exception as e:
            print(e)
            pass
    print('completado')
    print(lista)


@admin.register(LoadProfile)
class LoadProfileAdmin(admin.ModelAdmin):
    list_display = ('id_profile', 'arc_id',)
    actions = [update_profile]


@admin.register(LogSiebelOvPiano)
class LogSiebelOvPianoAdmin(admin.ModelAdmin):
    list_display = ('transaction', 'created', )
    search_fields = ('transaction__payu_transaction_id', 'transaction__subscription_id_str',)
    list_filter = (('transaction__subscription__start_date', DateTimeRangeFilter),)
    readonly_fields = (
        'log_request', 'log_response', 'transaction', 'fecha',
    )


@admin.register(LogUnsubscribePiano)
class LogUnsubscribePianoAdmin(admin.ModelAdmin):
    list_display = ('subscription_low', 'sent_to_siebel', 'siebel_request', 'siebel_response', )


@admin.register(ArcUnsubscribe)
class ArcUnsubscribeAdmin(admin.ModelAdmin):
    list_display = ('subscription', )
    list_filter = ('sent_to_siebel', )


@admin.register(LogRenovationPiano)
class LogRenovationPianoAdmin(ExportMixin, admin.ModelAdmin):
    resource_class = LogRenovationPianoResource
    list_display = ('transaction', 'get_id_transaction_payu', 'state', 'created', 'last_updated',)
    search_fields = ('transaction__payu_transaction_id', 'transaction__subscription_id_str', 'siebel_response',)
    readonly_fields = (
        'transaction', 'state',
    )

    def get_id_transaction_payu(self, obj):
        if obj.transaction:
            return obj.transaction.payu_transaction_id
        else:
            return ''


@admin.register(LogSiebelPaymentPiano)
class LogSiebelPaymentPianoAdmin(admin.ModelAdmin):
    list_display = ('transaction', 'get_id_transaction_payu', 'state',)
    search_fields = ('transaction__payu_transaction_id', 'transaction__subscription_id_str',)

    def get_id_transaction_payu(self, obj):
        if obj.transaction:
            return obj.transaction.payu_transaction_id
        else:
            return ''
